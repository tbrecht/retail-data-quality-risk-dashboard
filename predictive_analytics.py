import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from validation import load_config

CONFIG = load_config()

INPUT_FILE = CONFIG["cleaning_output_file"]
OUTPUT_FILE = CONFIG["predictive_output_file"]
SHEET_NAME = "Cleaned_Data"
COLUMN_CONFIG = CONFIG["columns"]


def normalize_score(series):
    series = pd.to_numeric(series, errors="coerce").fillna(0)

    if series.max() == series.min():
        return pd.Series([0] * len(series), index=series.index)

    return ((series - series.min()) / (series.max() - series.min())) * 100


def risk_level(score):
    if score >= 75:
        return "Critical"
    if score >= 50:
        return "High"
    if score >= 25:
        return "Moderate"
    return "Low"


def validate_required_columns(df):
    required_fields = ["sales", "category"]

    missing = []

    for field in required_fields:
        configured_column = COLUMN_CONFIG.get(field)

        if configured_column is None:
            missing.append({
                "configured_field": field,
                "problem": "Missing from config.json"
            })
        elif configured_column not in df.columns:
            missing.append({
                "configured_field": field,
                "problem": f"Column '{configured_column}' not found in cleaned dataset"
            })

    optional_fields = ["date", "customer"]

    for field in optional_fields:
        configured_column = COLUMN_CONFIG.get(field)

        if configured_column and configured_column not in df.columns:
            missing.append({
                "configured_field": field,
                "problem": f"Optional column '{configured_column}' not found in cleaned dataset"
            })

    if missing:
        missing_df = pd.DataFrame(missing)
        raise ValueError(
            "Column configuration issue found:\n"
            + missing_df.to_string(index=False)
        )


def calculate_outlier_burden(df, category_col, sales_col):
    values = pd.to_numeric(df[sales_col], errors="coerce").dropna()

    if values.empty:
        return pd.DataFrame(columns=[category_col, "outlier_count", "outlier_sales", "outlier_burden"])

    q1 = values.quantile(0.25)
    q3 = values.quantile(0.75)
    iqr = q3 - q1

    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr

    df = df.copy()

    df["is_outlier"] = (
        (pd.to_numeric(df[sales_col], errors="coerce") < lower_bound) |
        (pd.to_numeric(df[sales_col], errors="coerce") > upper_bound)
    )

    outliers = (
        df.groupby(category_col)
        .agg(
            outlier_count=("is_outlier", "sum"),
            total_records=(sales_col, "count"),
            total_sales=(sales_col, "sum")
        )
        .reset_index()
    )

    outlier_sales = (
        df[df["is_outlier"]]
        .groupby(category_col)[sales_col]
        .sum()
        .reset_index(name="outlier_sales")
    )

    outliers = outliers.merge(outlier_sales, on=category_col, how="left")
    outliers["outlier_sales"] = outliers["outlier_sales"].fillna(0)

    outliers["outlier_burden"] = (
        outliers["outlier_sales"] / outliers["total_sales"].replace(0, np.nan)
    ).replace([np.inf, -np.inf], 0).fillna(0)

    return outliers[[category_col, "outlier_count", "outlier_sales", "outlier_burden"]]


def apply_formatting_and_charts(output_file):
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(size=14)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 4, 65)

    if "Model_Insights" in wb.sheetnames:
        ws = wb["Model_Insights"]

        chart = BarChart()
        chart.title = "Top Segment Risk Scores"
        chart.y_axis.title = "Risk Score"
        chart.x_axis.title = "Segment"

        data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 11))
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 11))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 24

        ws.add_chart(chart, "H2")

    wb.save(output_file)


def run_predictive_analysis():
    if not Path(INPUT_FILE).exists():
        raise FileNotFoundError(f"{INPUT_FILE} not found. Run clean_retail_data.py first.")

    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)

    validate_required_columns(df)

    sales_col = COLUMN_CONFIG["sales"]
    category_col = COLUMN_CONFIG["category"]
    date_col = COLUMN_CONFIG.get("date")
    customer_col = COLUMN_CONFIG.get("customer")

    df[sales_col] = pd.to_numeric(df[sales_col], errors="coerce")
    df = df.dropna(subset=[sales_col, category_col])

    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        df = df.dropna(subset=[date_col])

    grouped = df.groupby(category_col)[sales_col].agg(
        total_sales="sum",
        average_sale="mean",
        sales_sd="std",
        transaction_count="count"
    ).reset_index()

    grouped["sales_sd"] = grouped["sales_sd"].fillna(0)

    grouped["sales_volatility"] = (
        grouped["sales_sd"] / grouped["average_sale"]
    ).replace([np.inf, -np.inf], 0).fillna(0)

    grouped["sales_share"] = grouped["total_sales"] / grouped["total_sales"].sum()

    if date_col:
        df = df.sort_values(date_col)
        midpoint = df[date_col].median()

        early = (
            df[df[date_col] <= midpoint]
            .groupby(category_col)[sales_col]
            .sum()
            .reset_index(name="early_sales")
        )

        recent = (
            df[df[date_col] > midpoint]
            .groupby(category_col)[sales_col]
            .sum()
            .reset_index(name="recent_sales")
        )

        grouped = grouped.merge(early, on=category_col, how="left")
        grouped = grouped.merge(recent, on=category_col, how="left")

        grouped["early_sales"] = grouped["early_sales"].fillna(0)
        grouped["recent_sales"] = grouped["recent_sales"].fillna(0)

        grouped["sales_momentum"] = (
            (grouped["recent_sales"] - grouped["early_sales"]) /
            grouped["early_sales"].replace(0, np.nan)
        ).replace([np.inf, -np.inf], np.nan).fillna(0)

        most_recent_date = df[date_col].max()
        df["days_old"] = (most_recent_date - df[date_col]).dt.days
        df["recency_weight"] = np.exp(-df["days_old"] / 60)

        recency_weighted_sales = (
            df.assign(weighted_sales=df[sales_col] * df["recency_weight"])
            .groupby(category_col)
            .agg(
                recency_weighted_sales=("weighted_sales", "sum"),
                recency_weight_total=("recency_weight", "sum")
            )
            .reset_index()
        )

        recency_weighted_sales["recency_weighted_avg_sales"] = (
            recency_weighted_sales["recency_weighted_sales"] /
            recency_weighted_sales["recency_weight_total"].replace(0, np.nan)
        ).replace([np.inf, -np.inf], 0).fillna(0)

        grouped = grouped.merge(
            recency_weighted_sales[
                [category_col, "recency_weighted_sales", "recency_weighted_avg_sales"]
            ],
            on=category_col,
            how="left"
        )

    else:
        grouped["early_sales"] = np.nan
        grouped["recent_sales"] = np.nan
        grouped["sales_momentum"] = 0
        grouped["recency_weighted_sales"] = grouped["total_sales"]
        grouped["recency_weighted_avg_sales"] = grouped["average_sale"]

    if customer_col:
        customer_sales = (
            df.groupby([category_col, customer_col])[sales_col]
            .sum()
            .reset_index()
        )

        top_customer_sales = (
            customer_sales
            .sort_values([category_col, sales_col], ascending=[True, False])
            .groupby(category_col)
            .head(1)
            .rename(columns={sales_col: "top_customer_sales"})
        )

        total_segment_sales = (
            df.groupby(category_col)[sales_col]
            .sum()
            .reset_index(name="segment_sales_for_customer_calc")
        )

        customer_concentration = top_customer_sales.merge(
            total_segment_sales,
            on=category_col,
            how="left"
        )

        customer_concentration["customer_concentration"] = (
            customer_concentration["top_customer_sales"] /
            customer_concentration["segment_sales_for_customer_calc"].replace(0, np.nan)
        ).replace([np.inf, -np.inf], 0).fillna(0)

        grouped = grouped.merge(
            customer_concentration[
                [category_col, customer_col, "top_customer_sales", "customer_concentration"]
            ],
            on=category_col,
            how="left"
        )

        grouped = grouped.rename(columns={customer_col: "top_customer"})

    else:
        grouped["top_customer"] = "Not available"
        grouped["top_customer_sales"] = np.nan
        grouped["customer_concentration"] = 0

    outlier_burden = calculate_outlier_burden(df, category_col, sales_col)
    grouped = grouped.merge(outlier_burden, on=category_col, how="left")

    grouped["outlier_count"] = grouped["outlier_count"].fillna(0)
    grouped["outlier_sales"] = grouped["outlier_sales"].fillna(0)
    grouped["outlier_burden"] = grouped["outlier_burden"].fillna(0)

    grouped["low_sales_score"] = 100 - normalize_score(grouped["total_sales"])
    grouped["volatility_score"] = normalize_score(grouped["sales_volatility"])
    grouped["decline_score"] = normalize_score(-grouped["sales_momentum"])
    grouped["low_volume_score"] = 100 - normalize_score(grouped["transaction_count"])
    grouped["customer_concentration_score"] = normalize_score(grouped["customer_concentration"])
    grouped["outlier_burden_score"] = normalize_score(grouped["outlier_burden"])

    grouped["data_quality_risk_score"] = normalize_score(
        grouped["outlier_burden_score"] + grouped["low_volume_score"]
    )

    grouped["business_trend_risk"] = grouped["decline_score"]

    grouped["operational_risk"] = (
        0.60 * grouped["volatility_score"] +
        0.40 * grouped["low_volume_score"]
    )

    grouped["customer_risk"] = grouped["customer_concentration_score"]

    grouped["performance_risk"] = grouped["low_sales_score"]

    grouped["data_reliability_risk"] = (
        0.65 * grouped["outlier_burden_score"] +
        0.35 * grouped["data_quality_risk_score"]
    )

    grouped["composite_risk_score"] = (
        0.25 * grouped["business_trend_risk"] +
        0.20 * grouped["operational_risk"] +
        0.15 * grouped["customer_risk"] +
        0.20 * grouped["performance_risk"] +
        0.20 * grouped["data_reliability_risk"]
    ).round(1)

    grouped["risk_level"] = grouped["composite_risk_score"].apply(risk_level)
    grouped = grouped.sort_values("composite_risk_score", ascending=False)

    model_insights = grouped[
        [
            category_col,
            "composite_risk_score",
            "risk_level",
            "total_sales",
            "sales_momentum",
            "sales_volatility",
            "customer_concentration",
            "outlier_burden"
        ]
    ].head(10).copy()

    model_insights["nontechnical_interpretation"] = model_insights.apply(
        lambda row: (
            f"{row[category_col]} is classified as {row['risk_level']} risk "
            f"with a risk score of {row['composite_risk_score']}. "
            f"This is a relative ranking score, not a probability. "
            f"The score considers recent sales movement, sales instability, total sales strength, "
            f"transaction volume, customer concentration, and outlier exposure."
        ),
        axis=1
    )

    risk_explainer = pd.DataFrame({
        "section": [
            "What the risk score means",
            "What the risk score does NOT mean",
            "Score range",
            "Low risk cutoff",
            "Moderate risk cutoff",
            "High risk cutoff",
            "Critical risk cutoff",
            "Updated risk score formula",
            "Business trend risk",
            "Operational risk",
            "Customer risk",
            "Performance risk",
            "Data reliability risk"
        ],
        "explanation": [
            "The composite risk score is a relative business risk ranking from 0 to 100. Higher scores indicate that a segment should receive more attention.",
            "The score is not a probability. A score of 60 does not mean there is a 60% chance that the segment will fail or decline.",
            "Scores range from 0 to 100, where 0 indicates the lowest relative concern and 100 indicates the highest relative concern within this dataset.",
            "0 to 24.9 = Low risk",
            "25 to 49.9 = Moderate risk",
            "50 to 74.9 = High risk",
            "75 to 100 = Critical risk",
            "Composite Risk Score = 25% business trend risk + 20% operational risk + 15% customer risk + 20% performance risk + 20% data reliability risk.",
            "Business trend risk measures whether recent sales are weakening compared with earlier sales.",
            "Operational risk combines sales volatility and low transaction volume.",
            "Customer risk measures whether a segment depends heavily on one customer.",
            "Performance risk measures whether a segment has weaker sales relative to other segments.",
            "Data reliability risk measures the degree to which outliers and low data support may reduce confidence in the segment."
        ]
    })

    executive_summary = pd.DataFrame({
        "section": [
            "Purpose",
            "Main result",
            "Risk score reminder",
            "Recommended action"
        ],
        "summary": [
            "This analysis identifies which business segments appear most at risk of future underperformance.",
            f"The highest-risk segment is {grouped.iloc[0][category_col]} with a risk score of {grouped.iloc[0]['composite_risk_score']} ({grouped.iloc[0]['risk_level']} risk).",
            "The risk score is a relative ranking score, not a probability. A score of 60 does not mean a 60% chance of decline.",
            "Review High and Critical risk segments first. These may require pricing review, marketing support, inventory review, customer concentration review, data quality review, or deeper operational investigation."
        ]
    })

    scientific_breakdown = pd.DataFrame({
        "topic": [
            "Model type",
            "Business objective",
            "Unit of analysis",
            "Outcome being estimated",
            "Configured columns used",
            "Features engineered",
            "Updated composite score formula",
            "Customer concentration",
            "Outlier burden",
            "Recency weighting",
            "Risk score interpretation",
            "Risk score cutoffs",
            "Why this approach was used",
            "Model limitations",
            "Data limitations",
            "Recommended next step"
        ],
        "details": [
            "Transparent composite risk scoring model using engineered sales-performance, customer, and data reliability features.",
            "Identify segments that may be at risk of future underperformance or require business review.",
            f"Each row in the model output represents one value of the configured category column: '{category_col}'.",
            "The model estimates relative business risk, not a guaranteed future outcome.",
            f"Sales column: {sales_col}; Category column: {category_col}; Date column: {date_col}; Customer column: {customer_col}.",
            "Total sales, average sale, transaction count, sales volatility, sales share, sales momentum, recency-weighted sales, customer concentration, outlier burden, and normalized risk component scores.",
            "Composite Risk Score = 25% business trend risk + 20% operational risk + 15% customer risk + 20% performance risk + 20% data reliability risk.",
            "Customer concentration measures the share of segment sales driven by the largest customer. Higher concentration may indicate dependency risk.",
            "Outlier burden measures the share of segment sales tied to statistical outlier transactions. Higher burden may indicate unstable or unreliable performance signals.",
            "Recent transactions are weighted more heavily using an exponential decay function. This helps the model give more attention to recent performance.",
            "The risk score is a relative ranking from 0 to 100. It is not a probability and should not be interpreted as percent likelihood of decline.",
            "Low: 0-24.9; Moderate: 25-49.9; High: 50-74.9; Critical: 75-100.",
            "A transparent scoring model was chosen because it is easier for nontechnical stakeholders to understand and audit than a black-box model.",
            "This is not a causal model. It does not prove why performance changed. It ranks segments by relative risk based on observed data patterns.",
            "Results depend on the quality, completeness, and time coverage of the cleaned dataset. Limited history, missing values, outliers, or incomplete customer data may affect reliability.",
            "Use this model as an early warning system, then investigate high-risk segments with additional business context."
        ]
    })

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        executive_summary.to_excel(writer, sheet_name="Executive_Summary", index=False)
        model_insights.to_excel(writer, sheet_name="Model_Insights", index=False)
        risk_explainer.to_excel(writer, sheet_name="Risk_Score_Explainer", index=False)
        scientific_breakdown.to_excel(writer, sheet_name="Scientific_Breakdown", index=False)
        grouped.to_excel(writer, sheet_name="Risk_Model_Output", index=False)

    apply_formatting_and_charts(OUTPUT_FILE)

    print("\nPredictive analytics complete.")
    print(f"Output saved to: {OUTPUT_FILE}")
    print("\nNontechnical summary:")
    print(
        f"The highest-risk segment is {grouped.iloc[0][category_col]} "
        f"with a score of {grouped.iloc[0]['composite_risk_score']} "
        f"({grouped.iloc[0]['risk_level']} risk)."
    )
    print("Reminder: the risk score is a relative ranking, not a probability.")


if __name__ == "__main__":
    run_predictive_analysis()